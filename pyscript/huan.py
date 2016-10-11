#!/usr/bin/env python
# coding:utf-8
import urllib
import requests
from openpyxl import load_workbook

wb = load_workbook(filename="huan.xlsx")
sheets = wb.get_sheet_names()

ws = wb.get_sheet_by_name(sheets[0])
columns = ws.columns
rows = ws.rows

def getBaikeUrl(name):
    url = "http://baike.baidu.com/search/word?word={0}".format(name.encode("utf-8"))
    #resp = requests.get(url, allow_redirects=False)
    resp = requests.get(url, allow_redirects=False)
    url = resp.headers.get("Location")
    return url

print getBaikeUrl(u"郝邵文")
#NUM = 502
#for i in xrange(2, NUM):
#    name = ws["B{0}".format(i)].value
#    url = getBaikeUrl(name)
#    if url and ("search" not in url):
#        ws["E{0}".format(i)].value = url
#        print "已完成第: {0}个.".format(i)
#
#wb.save(filename="huan.xlsx")


#n = 0
#content = []
#for row in rows:
#    n += 1
#    if not row[1].value:
#        print n
    #name = row[1].value
    #url  = getBaikeUrl(name)
    #if url:
    #    print name
    #    print url

